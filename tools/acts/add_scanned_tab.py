#!/usr/bin/env python3
"""Add/refresh the ``scanned_docx`` tab in FBR_Document_Inventory.xlsx.

Merges three sources:
  * ``/tmp/scan_inv.json``          -- per-PAGE scan census of the 91 Acts PDFs
                                       (which pages carry an image and < 200
                                       chars of text), produced by the sweep
  * ``reports/ocr-exclusions.md``   -- measured inter-engine agreement per file
  * the workbook's ``Full Inventory`` sheet -- the Ordinance and Rules documents,
                                       which this project has NOT measured

Writes one row per document that contains at least one scanned page, with the
processing path each one needs.  Re-running replaces the tab.
"""

from __future__ import annotations

import json
import os
import pathlib
import re
import sys

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if _ROOT not in sys.path:
    sys.path.insert(0, _ROOT)

XLSX = pathlib.Path(_ROOT) / "FBR_Document_Inventory.xlsx"
CENSUS = pathlib.Path("/tmp/scan_inv.json")
EXCL = pathlib.Path(_ROOT) / "reports" / "ocr-exclusions.md"
TAB = "scanned_docx"

# Decisions taken on this project: dual-engine OCR (RapidOCR + Tesseract) with a
# fidelity floor at 85% inter-engine agreement, and vision reading for pages
# below 95% because that is where every corruption class was found.
FLOOR = 85.0
VISION_AT = 95.0


def human(n: int) -> str:
    for unit in ("B", "KB", "MB"):
        if n < 1024 or unit == "MB":
            return f"{n:.1f} {unit}" if unit != "B" else f"{n} B"
        n /= 1024.0
    return f"{n:.1f} MB"


def measured() -> dict:
    """``{basename: (agreement, low_conf, verdict)}`` from the OCR sweep."""
    out = {}
    if not EXCL.exists():
        return out
    row = re.compile(r"^\|\s*`([^`]+)`\s*\|\s*(\d+)\s*\|\s*(\d+)\s*\|\s*(\d+)\s*\|"
                     r"\s*(\d+)\s*\|\s*([\d.]+)%\s*\|\s*([\d.]+)%\s*\|\s*(.+?)\s*\|")
    for line in EXCL.read_text().splitlines():
        m = row.match(line)
        if not m:
            continue
        out[pathlib.Path(m.group(1)).name] = {
            "ocr_pages": int(m.group(3)),
            "tokens": int(m.group(5)),
            "agreement": float(m.group(6)),
            "low_conf": float(m.group(7)),
            "verdict": "EXCLUDE" if "EXCLUDE" in m.group(8) else "admit",
        }
    return out


def path_for(scan_frac: float, meas: dict | None) -> tuple[str, str]:
    """(processing path, note) for one document."""
    if meas is None:
        return ("Needs measurement",
                "not yet measured by this project -- run scripts/ocr_review.py")
    a = meas["agreement"]
    if meas["verdict"] == "EXCLUDE":
        return ("VISION REQUIRED",
                f"below the {FLOOR:.0f}% floor ({a:.1f}%) -- not shipped today; "
                f"vision must read it before it can be admitted")
    if a < VISION_AT:
        return ("Vision recommended",
                f"admitted at {a:.1f}% but under {VISION_AT:.0f}%: this is where "
                f"the OCR engines corrupt enumerators and drop lines")
    return ("OCR sufficient",
            f"{a:.1f}% inter-engine agreement; no vision pass needed")


def main() -> int:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    if not CENSUS.exists():
        print(f"missing {CENSUS} -- run the page census first", file=sys.stderr)
        return 2
    census = json.load(CENSUS.open())
    meas = measured()

    wb = openpyxl.load_workbook(XLSX)
    inv = wb["Full Inventory"]
    irows = list(inv.iter_rows(values_only=True))
    ih = {str(h): i for i, h in enumerate(irows[0])}

    rows = []

    # --- Acts: measured per-page census -------------------------------------
    for r in census:
        pages, sp = r.get("pages") or 0, r.get("scanned_pages") or []
        if not sp:
            continue
        frac = len(sp) / pages if pages else 0
        m = meas.get(r["file"])
        kind = "Wholly scanned" if frac >= 0.9 else "Mixed (text + image pages)"
        p, note = path_for(frac, m)
        pl = ", ".join(str(x) for x in sp[:10]) + ("  …" if len(sp) > 10 else "")
        rows.append(["Acts", r["family"], r["file"], pages, len(sp),
                     f"{frac:.0%}", kind, pl, human(r.get("bytes") or 0),
                     f'{m["agreement"]:.2f}' if m else "",
                     f'{m["low_conf"]:.2f}' if m else "",
                     m["tokens"] if m else "",
                     m["verdict"] if m else "not measured", p, note])

    # --- Ordinance + Rules: from the inventory, NOT measured here -----------
    for r in irows[1:]:
        if not r or not r[0]:
            continue
        cat = str(r[ih["Category"]])
        if cat == "Acts" or "Scanned" not in str(r[ih["Content Type"]]):
            continue
        pages = int(r[ih["Pages"]] or 0)
        rows.append([cat, str(r[ih["Document Group"]]), str(r[ih["File Name"]]),
                     pages, pages, "100%", "Wholly scanned (per inventory)", "",
                     str(r[ih["Size"]]), "", "", "", "not measured",
                     "Needs measurement",
                     "outside this project's scope so far (Acts only); "
                     "same dual-engine + vision path applies"])

    rows.sort(key=lambda x: (x[0], -(x[4] or 0)))

    if TAB in wb.sheetnames:
        del wb[TAB]
    ws = wb.create_sheet(TAB)

    bold = Font(bold=True)
    title = Font(bold=True, size=13)
    hdr_fill = PatternFill("solid", fgColor="DDEBF7")
    warn_fill = PatternFill("solid", fgColor="FFC7CE")
    rec_fill = PatternFill("solid", fgColor="FFEB9C")
    ok_fill = PatternFill("solid", fgColor="E2EFDA")

    ws["A1"] = "Scanned documents — OCR / vision requirements"
    ws["A1"].font = title
    notes = [
        "Method: a page counts as SCANNED when it carries an image and yields "
        "under 200 characters of extractable text.",
        "Agreement % = inter-engine agreement between RapidOCR and Tesseract on "
        "the same rendered page. It is NOT either engine's self-reported "
        "confidence: on a degraded scan Tesseract reports ~95 on tokens it got "
        "wrong, and only a second recogniser exposes that.",
        f"Fidelity floor = {FLOOR:.0f}% mean agreement AND <=15% low-confidence "
        f"tokens. Below the floor a document is NOT shipped.",
        f"Vision is used where agreement < {VISION_AT:.0f}% — measured to be "
        "where the engines mangle enumerators ((a)->(al, (c)->(c}) and drop whole "
        "lines, including one operative heading.",
        "Acts rows are measured by this project. Ordinance and Rules rows are "
        "carried from the Full Inventory sheet and have NOT been measured here.",
        "IMPORTANT: the Full Inventory marks only 2 documents 'Mixed' — the "
        "per-page census below finds many more text-layer documents that contain "
        "a few image-only pages. Those pages are invisible to a per-document "
        "check and to both sides of a text-conservation audit.",
    ]
    r = 3
    for n in notes:
        ws.cell(row=r, column=1, value="• " + n).alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=15)
        ws.row_dimensions[r].height = 28
        r += 1

    r += 1
    head = ["Category", "Document Group", "File Name", "Total Pages",
            "Scanned Pages", "Scanned %", "Scan Type", "Scanned Page Numbers",
            "Size", "Agreement %", "Low-conf %", "OCR Tokens", "Verdict",
            "Processing Path", "Notes"]
    for c, h in enumerate(head, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = bold
        cell.fill = hdr_fill
        cell.alignment = Alignment(wrap_text=True, vertical="bottom")
    ws.freeze_panes = ws.cell(row=r + 1, column=1)
    hdr_row = r

    for row in rows:
        r += 1
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            if c in (8, 15):
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        p = row[13]
        fill = (warn_fill if p == "VISION REQUIRED" else
                rec_fill if p in ("Vision recommended", "Needs measurement") else
                ok_fill)
        ws.cell(row=r, column=14).fill = fill
        ws.cell(row=r, column=13).fill = fill

    # summary under the table
    r += 2
    ws.cell(row=r, column=1, value="Summary").font = bold
    acts = [x for x in rows if x[0] == "Acts"]
    counts = {
        "Documents with at least one scanned page": len(rows),
        "  of which Acts (measured here)": len(acts),
        "  of which Ordinance / Rules (not measured here)": len(rows) - len(acts),
        "Wholly scanned": sum(1 for x in rows if x[6].startswith("Wholly")),
        "Mixed text + image": sum(1 for x in rows if x[6].startswith("Mixed")),
        "Total scanned pages (Acts, measured)": sum(x[4] for x in acts),
        "OCR sufficient (>= 95% agreement)": sum(1 for x in rows if x[13] == "OCR sufficient"),
        "Vision recommended (85-95%)": sum(1 for x in rows if x[13] == "Vision recommended"),
        "VISION REQUIRED (below floor, not shipped)": sum(1 for x in rows if x[13] == "VISION REQUIRED"),
        "Needs measurement": sum(1 for x in rows if x[13] == "Needs measurement"),
    }
    for k, v in counts.items():
        r += 1
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=4, value=v).font = bold

    for col, w in zip("ABCDEFGHIJKLMNO",
                      (11, 30, 52, 11, 13, 10, 26, 22, 10, 12, 11, 11, 13, 20, 60)):
        ws.column_dimensions[col].width = w

    wb.save(XLSX)
    print(f"wrote '{TAB}' to {XLSX.name}: {len(rows)} rows "
          f"({len(acts)} Acts measured, {len(rows) - len(acts)} carried over)")
    for k, v in counts.items():
        print(f"   {k:52s} {v}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
